# _*_ coding: UTF-8 _*_

import os
import re
import sys
import zipfile
import openpyxl as xl
import win32com.client as win32
import pdfplumber
import logging

from ui.ui_EMC import *
from ui.ui_docx2pdf import Ui_Docx2PdfWin
from PySide6.QtWidgets import QProgressBar, QMessageBox, QCheckBox, QDialog, QFileDialog
from PySide6.QtCore import QThread, Signal, Slot
from qt_material import QtStyleTools, apply_stylesheet

# 日志配置
logName = "DealPdf"
logger = logging.getLogger(logName)
logger.setLevel(logging.INFO)
handler = logging.FileHandler(logName + ".log")
handler.setFormatter(logging.Formatter("%(asctime)s - %(levelname)s - %(message)s"))
handler.setLevel(logging.INFO)
logger.addHandler(handler)


class ConvertThread(QThread):
    docx_count = Signal(int)
    docx_curr = Signal(int)
    docx_end = Signal()

    def __init__(self, docx_path=None):
        super().__init__()
        self.docx_path = docx_path

    def run(self):
        if self.docx_path is None:
            return
        else:
            docx_path = self.docx_path
        try:
            word = win32.Dispatch("Word.Application")
            word.Visible = False
        except:
            print("Please install Microsoft Word 2010 or later.")
            logger.error("Please install Microsoft Word 2010 or later.")
            QMessageBox.warning(
                self, "Error", "Please install Microsoft Word 2010 or later."
            )
            return
        docx_path = os.path.abspath(docx_path)
        if os.path.isdir(docx_path):
            source_files = [f for f in os.listdir(docx_path) if f.endswith(".docx")]
        else:
            source_files = [docx_path]
        file_count = len(source_files)
        self.docx_count.emit(file_count)
        curr = 0
        for file in source_files:
            docx_file = os.path.join(docx_path, file)
            pdf_file = os.path.join(docx_path, file.replace(".docx", ".pdf"))
            try:
                doc = word.Documents.Open(os.path.abspath(docx_file))
                doc.SaveAs(os.path.abspath(pdf_file), FileFormat=17)
                doc.Close(SaveChanges=0)
            except Exception as e:
                logger.error(f"Error converting {docx_file} to PDF: {e}")
            curr = curr + 1
            self.docx_curr.emit(curr)
        self.docx_end.emit()
        logger.info(f"Converted {curr} success/{file_count-curr} fail files to PDF.")
        word.Quit()


class Docx2PdfWindow(QMainWindow, Ui_Docx2PdfWin, QtStyleTools):
    def __init__(self, parent=None, rootpath=None):
        super().__init__(parent)
        self.setupUi(self)
        self.rootpath = (
            rootpath if rootpath else os.path.join(os.path.expanduser("~"), "Documents")
        )
        self.docx2pdf_thread = ConvertThread()

        self.btn_docxdir.clicked.connect(lambda: self.select_docx_path(is_dir=True))
        self.button_docxfile.clicked.connect(
            lambda: self.select_docx_path(is_dir=False)
        )
        self.btn_docx2pdf.clicked.connect(self.docx2pdf_thread.run)
        self.docx2pdf_thread.docx_count.connect(self.createStatusBar)
        self.docx2pdf_thread.docx_curr.connect(self.updateStatusBar)
        self.docx2pdf_thread.docx_end.connect(self.endStatus)
        self.docx2pdf_thread.start()

    def select_docx_path(self, is_dir=True):
        if is_dir:
            path = QFileDialog.getExistingDirectory(
                self,
                "选择DOCX文件所在目录",
                dir=self.rootpath,
            )
            self.rootpath = os.path.split(path)[0]
            if path:
                self.textEdit_docx.setText(path)
                self.docx2pdf_thread.docx_path = path
        else:
            path = QFileDialog.getOpenFileName(
                self,
                "选择DOCX文件",
                filter="DOCX文件 (*.docx)",
                dir=self.rootpath,
            )
            self.rootpath = os.path.split(os.path.dirname(path[0]))[0]
            if path:
                self.textEdit_docx.setText(path[0])
                self.docx2pdf_thread.docx_path = path[0]

    @Slot(int)
    def createStatusBar(self, file_count):
        self.file_count = file_count
        # 创建状态栏进度条
        self.statusBar_docx.showMessage("正在转换DOCX文件...")
        self.pgBar = QProgressBar(self.statusBar_docx)
        self.pgBar.setMinimum(0)
        self.pgBar.setMaximum(file_count)
        self.pgBar.setValue(0)
        self.statusBar_docx.addPermanentWidget(self.pgBar)

    @Slot(int)
    def updateStatusBar(self, value):
        self.curr = value
        self.pgBar.setValue(value)
        self.statusBar_docx.showMessage(f"正在转换DOCX文件...{value}/{self.file_count}")

    @Slot()
    def endStatus(self):
        self.statusBar_docx.removeWidget(self.pgBar)
        self.statusBar_docx.showMessage(
            f"转换DOCX文件:{self.curr}成功/{self.file_count-self.curr}失败"
        )

    def closeEvent(self, event):
        self.deleteLater()
        return super().closeEvent(event)


class EMIWindow(QMainWindow, Ui_MainWindow, QtStyleTools):
    def __init__(self, parent=None):
        super(EMIWindow, self).__init__(parent)
        self.setupUi(self)

        self.row_sn_ = 44  # 序列号所在行
        self.col_sn_ = 4  # 序列号所在列
        self.col_vol_ = 11  # 电压所在列
        self.col_line_ = 13  # 线性所在列
        self.col_load_ = 14  # 负荷所在列
        self.row_end_ = 107  # 结束行
        self.col_end_ = 27  # 结束列
        self.rootpath = None  # 根目录

        self.settings_value = {
            "NeedZip": True,  # 是否需要压缩
            "AddZip": True,  # 是否嵌入压缩包到excel文件
            "CloseExcel": False,  # 是否在处理完毕后关闭excel程序
        }
        self.settings = {}  # 设置
        self.windows: list[QWidget] = []  # 打开的窗口
        self.log_path = None  # 日志文件路径

        self.btn_exit.clicked.connect(self.close)
        self.btn_deal.clicked.connect(self.deal_pdf)
        self.btn_pathName.clicked.connect(self.select_path)
        self.btn_pathTemp.clicked.connect(self.select_path)
        self.actionsetting.triggered.connect(self.show_setting)
        self.actionabout.triggered.connect(self.show_about)
        self.actionhelpdoc.triggered.connect(self.show_helpdoc)
        self.actionlog.triggered.connect(self.show_log)

        self.docx2pdf_win = Docx2PdfWindow(rootpath=self.rootpath)
        self.actiondocx2pdf.triggered.connect(self.docx2pdf_win.show)
        self.windows.append(self.docx2pdf_win)
        # self.docx2pdf_thread.start()

    def select_path(self):
        if self.sender() == self.btn_pathName:
            path = QFileDialog.getExistingDirectory(
                self,
                "选择原始PDF文件所在目录",
                dir=(
                    self.rootpath
                    if self.rootpath
                    else os.path.join(os.path.expanduser("~"), "Documents")
                ),
            )
            self.rootpath = os.path.split(path)[0]
            if path:
                self.textEdit_name.setText(path)
        elif self.sender() == self.btn_pathTemp:
            path = QFileDialog.getOpenFileName(
                self,
                "选择模板",
                filter="Excel文件 (*.xlsx)",
                dir=(
                    self.rootpath
                    if self.rootpath
                    else os.path.join(os.path.expanduser("~"), "Documents")
                ),
            )
            self.rootpath = os.path.split(os.path.dirname(path[0]))[0]
            if path:
                self.textEdit_tempFile.setText(path[0])

    def closeEvent(self, event):
        # 关闭所有窗口
        for window in self.windows:
            window.close()
        return super().closeEvent(event)

    def update_Setting(self):
        for key, _ in self.settings.items():
            self.settings_value[key] = self.settings[key].isChecked()
        logger.info(f"Setting updated: {self.settings_value}")

    def show_setting(self):
        self.setting_win = QWidget()
        self.setting_win.setWindowTitle("设置")
        icon = QIcon(":/emipdf/acbel -1.jpg")
        self.setting_win.setWindowIcon(icon)
        self.setting_win.resize(300, 300)
        self.setting_win.setStyleSheet("QLabel { font-size: 15px; }")

        layout = QVBoxLayout()

        layoutH_1 = QHBoxLayout()
        label_NeedZip = QLabel(
            text="生成压缩包:", alignment=Qt.AlignRight | Qt.AlignVCenter
        )
        label_NeedZip.setToolTip("是否生成压缩包，压缩包中包含原始PDF文件。")
        layoutH_1.addWidget(label_NeedZip)
        self.NeedZip_checkbox = QCheckBox()
        self.NeedZip_checkbox.setChecked(self.settings_value["NeedZip"])
        layoutH_1.addWidget(self.NeedZip_checkbox)
        layout.addLayout(layoutH_1)
        self.settings["NeedZip"] = self.NeedZip_checkbox

        layoutH_2 = QHBoxLayout()
        label_addZip = QLabel(
            text="嵌入压缩包:", alignment=Qt.AlignRight | Qt.AlignVCenter
        )
        label_addZip.setToolTip("是否嵌入压缩包到excel文件。")
        layoutH_2.addWidget(label_addZip)
        self.addZip_checkbox = QCheckBox()
        self.addZip_checkbox.setChecked(self.settings_value["AddZip"])
        layoutH_2.addWidget(self.addZip_checkbox)
        layout.addLayout(layoutH_2)
        self.settings["AddZip"] = self.addZip_checkbox

        layoutH_3 = QHBoxLayout()
        label_closeExcel = QLabel(
            text="执行完毕关闭Excel程序:", alignment=Qt.AlignRight | Qt.AlignVCenter
        )
        label_closeExcel.setToolTip("是否在处理完毕后关闭excel程序，默认保留。")
        layoutH_3.addWidget(label_closeExcel)
        self.closeExcel_checkbox = QCheckBox()
        self.closeExcel_checkbox.setChecked(self.settings_value["CloseExcel"])
        layoutH_3.addWidget(self.closeExcel_checkbox)
        layout.addLayout(layoutH_3)
        self.settings["CloseExcel"] = self.closeExcel_checkbox

        layoutBTN = QHBoxLayout()
        button_Save = QPushButton(text="保存")
        button_Save.clicked.connect(self.update_Setting)
        button_Save.clicked.connect(self.setting_win.close)
        layoutBTN.addWidget(button_Save)

        button_Cancel = QPushButton(text="取消")
        button_Cancel.clicked.connect(self.setting_win.close)
        layoutBTN.addWidget(button_Cancel)

        layout.addLayout(layoutBTN)
        self.setting_win.setLayout(layout)
        self.setting_win.show()
        self.windows.append(self.setting_win)

    def export_log(self):
        export_path = QFileDialog.getSaveFileName(
            self, "导出日志", filter="Log Files (*.log)"
        )
        with open(self.log_path, "r") as sf:
            with open(export_path, "w") as f:
                f.write(sf.read())

    def clear_log(self):
        open(self.log_path, "w").close()
        self.Text_log.clear()
        logger.info("Log cleared.")

    def show_log(self):
        self.log_win = QDialog()
        self.log_win.setWindowTitle("日志")
        icon = QIcon()
        icon.addFile(":/emipdf/acbel -1.jpg", QSize(), QIcon.Normal, QIcon.Off)
        self.log_win.setWindowIcon(icon)
        self.log_win.resize(800, 600)
        self.log_win.setStyleSheet("QLabel { font-size: 15px; }")
        self.log_win.setLayout(QVBoxLayout())

        self.log_path = os.path.join(os.getcwd(), logName + ".log")

        self.Text_log = QTextEdit(
            text=open(self.log_path, "r").read(), alignment=Qt.AlignLeft, readOnly=True
        )
        self.Text_log.setLineWrapMode(QTextEdit.WidgetWidth)
        self.log_win.layout().addWidget(self.Text_log)

        layout_btn = QHBoxLayout()

        btn_clear = QPushButton(text="清 空 日 志")
        btn_clear.clicked.connect(self.clear_log)
        layout_btn.addWidget(btn_clear)

        btn_export = QPushButton(text="导 出 日 志")
        btn_export.clicked.connect(self.export_log)
        layout_btn.addWidget(btn_export)

        self.log_win.layout().addLayout(layout_btn)

        self.log_win.show()
        self.windows.append(self.log_win)

    def show_about(self):
        self.about_win = QWidget()
        self.about_win.setWindowTitle("关于")
        icon = QIcon()
        icon.addFile(":/emipdf/acbel -1.jpg", QSize(), QIcon.Normal, QIcon.Off)
        self.about_win.setWindowIcon(icon)
        self.about_win.resize(300, 200)
        self.about_win.setStyleSheet("QLabel { font-size: 15px; }")
        self.about_win.setLayout(QVBoxLayout())

        label_about = QLabel(
            text="EMI-Report\n\n版本：1.2.0\n\n作者：Lucas Li\n\n邮箱：Lucas_Li@acbel.com",
            alignment=Qt.AlignCenter,
        )
        label_about.setWordWrap(True)
        self.about_win.layout().addWidget(label_about)

        self.about_win.show()
        self.windows.append(self.about_win)

    def show_helpdoc(self):
        self.helpdoc_win = QWidget()
        self.helpdoc_win.setWindowTitle("帮助文档")
        icon = QIcon()
        icon.addFile(":/emipdf/acbel -1.jpg", QSize(), QIcon.Normal, QIcon.Off)
        self.helpdoc_win.setWindowIcon(icon)
        self.helpdoc_win.resize(800, 600)
        self.helpdoc_win.setStyleSheet("QLabel { font-size: 15px; }")
        self.helpdoc_win.setLayout(QVBoxLayout())

        label_helpdoc = QLabel(
            text="EMI-Report\n\n待补充...\n\n", alignment=Qt.AlignLeft
        )
        label_helpdoc.setWordWrap(True)
        self.helpdoc_win.layout().addWidget(label_helpdoc)

        self.helpdoc_win.show()
        self.windows.append(self.helpdoc_win)

    def find_pdf(self, directory):
        source_files = [f for f in os.listdir(directory) if f.endswith(".pdf")]
        res = {}
        file_count = len(source_files)

        # 创建状态栏进度条
        self.statusBar().showMessage("正在处理文件...")
        self.progressBar = QProgressBar(self.statusBar_main)
        self.progressBar.setMinimum(0)
        self.progressBar.setMaximum(file_count)
        self.progressBar.setValue(0)
        self.statusBar().addPermanentWidget(self.progressBar)

        for i, source_file in enumerate(source_files):
            logger.info(f"FileName: {source_file}")
            try:
                res[source_file] = self.open_pdf(os.path.join(directory, source_file))
            except Exception as e:
                logger.error(f"Error processing {source_file}: {e}")
                continue
            value = self.progressBar.value() + 1
            self.progressBar.setValue(value)
            self.statusBar().showMessage(f"正在处理文件...{value}/{file_count}")

        self.statusBar().removeWidget(self.progressBar)

        return res

    def open_pdf(self, path_pdf):
        # 读取pdf文件，提取数据
        with pdfplumber.open(path_pdf) as pdf:
            text = pdf.pages[0].extract_text() + "\n" + pdf.pages[1].extract_text()
            datas = [item for item in text.split("\n") if item]
            return self.deal_datas(datas)

    def deal_pdf(self):
        directory = self.textEdit_name.toPlainText()
        loadqty = self.textEdit_loadQTY.toPlainText()
        tmpFile = self.textEdit_tempFile.toPlainText()

        loadqty = re.match(r"\d+", loadqty).group() if loadqty else "3"

        # 检查是否存在模板文件，默认3个负载
        if not tmpFile or tmpFile == "":
            tmpFile = f"2.1 Conducted EMI Measurement_{loadqty}.xlsx"
            if loadqty not in ["3", "4"]:
                logger.warning(f"Unexpected load quantity: {loadqty}, defaulting to 3.")
                tmpFile = "2.1 Conducted EMI Measurement_3.xlsx"
            tmpFile = "template/" + tmpFile

        try:
            wb = xl.load_workbook(tmpFile)
        except FileNotFoundError:
            logger.warning(f"No template file found: {tmpFile}")
            QMessageBox.warning(self, "Warning", f"No template file found: {tmpFile}")
            return
        except Exception as e:
            logger.error(f"Error opening template file: {e}")
            QMessageBox.warning(self, "Error", f"Error opening template file: {e}")
            return

        ws_setup = wb["Setup"]

        self.row_sn_ = ws_setup.cell(row=1, column=2).value
        self.col_sn_ = ws_setup.cell(row=2, column=2).value
        self.col_vol_ = ws_setup.cell(row=3, column=2).value
        self.col_line_ = ws_setup.cell(row=4, column=2).value
        self.col_load_ = ws_setup.cell(row=5, column=2).value
        self.row_end_ = ws_setup.cell(row=7, column=2).value
        self.col_end_ = ws_setup.cell(row=8, column=2).value

        ws = wb["Conducted EMI"]

        if not os.path.exists(directory):
            logger.warning(f"No such directory: {directory}")
            QMessageBox.warning(self, "Warning", f"No such directory: {directory}")
            return

        res = self.find_pdf(directory)

        for sub in ["\\", "/", "\\\\"]:
            if directory.endswith(sub):
                directory = directory[:-1]
            if directory.find(sub) != -1:
                root_dir = directory
                directory = directory.split(sub)[-1]
                root_dir = root_dir.replace(directory, "") + sub
                break

        os.chdir(root_dir)  # 改变工作目录

        # 写入单体型号
        tmpp = directory.split("-")
        if len(tmpp) > 1:
            ws["F5"] = uutname = tmpp[0] + "-" + tmpp[1]
        else:
            ws["F5"] = directory

        num_uut = len(set([it.split("-")[0] for it in res.keys()]))

        if num_uut > 5:
            logger.warning(f"Number of Load is too many! {num_uut}")
            QMessageBox.warning(
                self, "Warning", f"Number of Load is too many! {num_uut}"
            )
            return

        # 处理合并单元格
        merged_cells = list(ws.merged_cells)
        for cell in merged_cells:
            if cell.min_row == self.row_sn_ + int(loadqty) * 4 * num_uut:
                ws.unmerge_cells(cell.coord)

        # 删去多余的行
        ws.delete_rows(
            self.row_sn_ + int(loadqty) * 4 * num_uut, int(loadqty) * 4 * (5 - num_uut)
        )
        self.row_end_ = ws.max_row - 4
        logger.info(f"Max row after deletion: {self.row_end_}")

        # 写入数据
        line_dict = {"L1": "Line", "N": "Neutral"}
        for item_key, datas in res.items():
            row_sn = self.row_sn_
            while (
                ws.cell(row=row_sn, column=self.col_sn_).value is not None
                and ws.cell(row=row_sn, column=self.col_sn_).value != datas["Serial"]
            ):
                row_sn += int(loadqty) * 4
            if ws.cell(row=row_sn, column=self.col_sn_).value is None:
                ws.cell(row=row_sn, column=self.col_sn_).value = datas["Serial"]

            row_vol = row_sn + (
                int(loadqty) * 2
                if ws.cell(row=row_sn, column=self.col_vol_).value
                != int(re.match(r"\d+", datas["Power"]).group())
                else 0
            )
            row_load = row_vol
            while row_load < row_vol + (int(loadqty) * 2) and int(
                ws.cell(row=row_load, column=self.col_load_).value * 100
            ) != int(re.match(r"\d+", datas["Load"]).group()):
                row_load += 2

            data = datas["datas"][0]
            row_line = (
                row_load + 1
                if ws.cell(row=row_load, column=self.col_line_).value
                != line_dict.get(data[9], "Unknown")
                else row_load
            )
            # 写入数据
            ws.cell(row=row_line, column=self.col_load_ + 1).value = data[0]
            ws.cell(row=row_line, column=self.col_load_ + 2).value = data[1]
            ws.cell(row=row_line, column=self.col_load_ + 3).value = data[3]
            ws.cell(row=row_line, column=self.col_load_ + 4).value = data[6]
            ws.cell(row=row_line, column=self.col_load_ + 5).value = "-"
            ws.cell(row=row_line, column=self.col_load_ + 6).value = data[2]
            ws.cell(row=row_line, column=self.col_load_ + 7).value = data[5]

        # 保存文件
        SaveFile = re.sub(r"_\d\.xlsx", f"_{uutname}.xlsx", os.path.basename(tmpFile))
        wb.remove(ws_setup)
        wb.save(SaveFile)
        print("Done!")
        if self.settings_value["NeedZip"]:
            try:
                self.zip_folder(directory, uutname)
                if self.settings_value["AddZip"]:
                    self.addZipToExcel(uutname, SaveFile)
            except Exception as e:
                logger.error(f"Error during zipping or embedding: {e}")
                QMessageBox.critical(
                    self, "Error", f"Error during zipping or embedding: {e}"
                )

        self.show_done(SaveFile)

    def deal_datas(self, datas):
        res = {}
        for data in datas[:20]:
            tmp = data.split(" ")
            if tmp[0] == "Serial":
                res["Serial"] = tmp[-1]
            elif tmp[0] == "Power":
                res["Power"] = tmp[-1]
            elif tmp[0] == "Load":
                res["Load"] = tmp[-1]
        resdatas = []
        for it in datas[15:]:
            data = it.split(" ")
            if len(data) > 10:
                resdatas.append(data)
        res["datas"] = resdatas
        res_avg = sorted(res["datas"], key=lambda x: float(x[7]), reverse=False)
        # 对pk列升序排列
        res_pk = sorted(res["datas"], key=lambda x: float(x[4]), reverse=False)
        # 取avg和pk中最小的
        res["datas"] = (
            res_avg if float(res_avg[0][7]) <= float(res_pk[0][4]) else res_pk
        )
        return res

    def addZipToExcel(self, zipname, SaveFile):
        try:
            excel = win32.Dispatch("Excel.Application")
            excel.Visible = True
        except:
            print("Please install Microsoft Word 2010 or later.")
            return
        try:
            wb = excel.Workbooks.Open(os.path.abspath(SaveFile))
            ws = wb.Worksheets("Conducted EMI")
            if ws.ProtectContents:
                ws.Unprotect()
            ws.Cells(self.row_sn_, self.col_end_ - 2).Select()
            ws.OLEObjects().Add(
                ClassType=None,
                Filename=os.path.abspath(f"{zipname}.zip"),
                Link=False,
                DisplayAsIcon=True,
            )
            # xlEdgeButtom = 9; xlMedium = -4138; xlContinuous = 1
            ws.Range(
                ws.Cells(self.row_end_, 2), ws.Cells(self.row_end_, self.col_end_ - 1)
            ).Select()
            ws.Range(
                ws.Cells(self.row_end_, 2), ws.Cells(self.row_end_, self.col_end_ - 1)
            ).Borders(9).LineStyle = 1
            ws.Range(
                ws.Cells(self.row_end_, 2), ws.Cells(self.row_end_, self.col_end_ - 1)
            ).Borders(9).Weight = -4138
            wb.Save()
            if self.settings_value["CloseExcel"]:
                excel.Quit()
        except Exception as e:
            logger.error(f"Error adding zip to Excel: {e}")
            QMessageBox.critical(self, "Error", f"Error adding zip to Excel: {e}")

    def zip_folder(self, directory, zipname):
        print("Zipping folder...")
        try:
            file_count = 0
            for rootPath, _, files in os.walk(directory):
                for file in files:
                    if file.endswith(".pdf"):
                        file_count += 1

            # 创建状态栏进度条
            self.statusBar().showMessage("正在压缩文件...")
            self.progressBar = QProgressBar(self.statusBar_main)
            self.progressBar.setMinimum(0)
            self.progressBar.setMaximum(file_count)
            self.progressBar.setValue(0)
            self.statusBar().addPermanentWidget(self.progressBar)

            print(f"Zipping {file_count} files...")
            with zipfile.ZipFile(f"{zipname}.zip", "w", zipfile.ZIP_LZMA) as HarZip:
                for rootPath, _, files in os.walk(directory):
                    for file in files:
                        if file.endswith(".pdf") == False:
                            continue
                        file_path = os.path.join(rootPath, file)
                        arcname = os.path.relpath(file_path, directory)
                        HarZip.write(file_path, arcname=arcname)

                        value = self.progressBar.value() + 1
                        self.progressBar.setValue(value)
                        self.statusBar().showMessage(
                            f"正在压缩文件...{value}/{file_count}"
                        )

            self.statusBar().removeWidget(self.progressBar)
            self.progressBar.deleteLater()
            self.statusBar().showMessage("压缩完成！")
            print("Done!")
        except Exception as e:
            logger.error(f"Error zipping folder: {e}")

    def show_done(self, SaveFile):
        self.done_win = QWidget()
        self.done_win.setWindowTitle("完成!")
        icon = QIcon(":/emipdf/acbel -1.jpg")
        self.done_win.setWindowIcon(icon)
        self.done_win.resize(600, 200)
        self.done_win.setLayout(QVBoxLayout())

        label1 = QLabel(
            f'文件已保存为 "{os.path.abspath(SaveFile)}"', alignment=Qt.AlignLeft
        )
        self.done_win.layout().addWidget(label1)

        label2_text = (
            "您还需要进行下列步骤以完成报告：\n"
            "\t1. 填入日期，测试者等表头信息，注意是MP还是MVT.\n"
            "\t2. 千万注意限值标准是否对应，默认Class B.\n"
            "\t3. 插入测试图片和压缩包.\n"
            "\t4. 整理表格，如果有多余项请按需求删除."
        )
        label2 = QLabel(text=label2_text, alignment=Qt.AlignLeft)
        label2.setWordWrap(True)
        self.done_win.layout().addWidget(label2)

        label3 = QLabel(
            "---------------请确认以上步骤都已完成，按OK退出。---------------",
            alignment=Qt.AlignLeft,
        )
        self.done_win.layout().addWidget(label3)

        ok_button = QPushButton("OK")
        ok_button.clicked.connect(self.done_win.hide)
        self.done_win.layout().addWidget(ok_button)

        self.done_win.show()
        self.windows.append(self.done_win)


if __name__ == "__main__":
    app = QApplication()
    win = EMIWindow()

    apply_stylesheet(app, theme="dark_teal.xml")
    win.show()
    sys.exit(app.exec())

    # win.addZipToExcel("FSC048-4C0G", "2.1 Conducted EMI Measurement.xlsx")
